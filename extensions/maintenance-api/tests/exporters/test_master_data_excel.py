from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from app.core.exceptions import BusinessValidationError
from app.models.catalog import SparePart
from app.models.supplier import Supplier
from openpyxl import load_workbook

SPARE_PART_HEADERS = [
    "器材编码",
    "器材名称",
    "规格型号",
    "类别",
    "单位",
    "制造商",
    "物料编码",
    "国军标",
    "保质期_月",
    "是否序列化",
    "是否可修复",
    "是否关键",
    "默认服务水平",
    "说明",
    "是否启用",
]


def _exporter():
    try:
        from app.exporters.master_data_excel import (
            MasterDataExcelExporter,
        )
    except ModuleNotFoundError:
        pytest.fail(
            "MASTER_DATA_EXCEL_EXPORTER_NOT_AVAILABLE"
        )

    return MasterDataExcelExporter()


def _rows_as_dicts(worksheet) -> list[dict[str, object]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    return [
        dict(zip(headers, row, strict=True))
        for row in rows
    ]


def test_export_respects_tenant_filter_sort_and_public_columns(
    session,
):
    session.add_all(
        [
            SparePart(
                tenant_id="tenant-a",
                code="SP-002",
                name="Tenant A Critical Two",
                unit="件",
                is_critical=True,
            ),
            SparePart(
                tenant_id="tenant-a",
                code="SP-001",
                name="Tenant A Critical One",
                unit="件",
                is_critical=True,
            ),
            SparePart(
                tenant_id="tenant-a",
                code="SP-003",
                name="Tenant A Noncritical",
                unit="件",
                is_critical=False,
            ),
            SparePart(
                tenant_id="tenant-b",
                code="SP-001",
                name="Tenant B Shared Code",
                unit="件",
                is_critical=True,
            ),
        ]
    )
    session.commit()

    content = _exporter().export(
        session,
        tenant_id="tenant-a",
        resource_key="spare-parts",
        filters={
            "is_critical": True,
            "sort_by": "code",
            "sort_order": "asc",
        },
    )

    workbook = load_workbook(
        BytesIO(content),
        data_only=False,
    )
    assert workbook.sheetnames == ["维修器材"]

    worksheet = workbook["维修器材"]
    headers = [
        cell.value
        for cell in worksheet[1]
    ]
    rows = _rows_as_dicts(worksheet)

    assert headers == SPARE_PART_HEADERS
    assert [
        row["器材编码"]
        for row in rows
    ] == ["SP-001", "SP-002"]
    assert {
        row["器材名称"]
        for row in rows
    } == {
        "Tenant A Critical One",
        "Tenant A Critical Two",
    }
    assert {
        row["是否关键"]
        for row in rows
    } == {True}
    assert all(
        "租户" not in str(header)
        for header in headers
    )
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref
    assert worksheet["A1"].font.bold is True


def test_export_escapes_excel_formula_prefixes(
    session,
):
    session.add(
        Supplier(
            tenant_id="tenant-a",
            code="FORMULA",
            name=(
                '=HYPERLINK("https://example.invalid")'
            ),
        )
    )
    session.commit()

    content = _exporter().export(
        session,
        tenant_id="tenant-a",
        resource_key="suppliers",
        filters={"keyword": "FORMULA"},
    )

    workbook = load_workbook(
        BytesIO(content),
        data_only=False,
    )
    assert workbook.sheetnames == ["供应商"]
    assert (
        workbook["供应商"]["B2"].value
        == (
            "'=HYPERLINK("
            '"https://example.invalid")'
        )
    )


def test_export_rejects_unknown_sort_field(
    session,
):
    with pytest.raises(BusinessValidationError):
        _exporter().export(
            session,
            tenant_id="tenant-a",
            resource_key="spare-parts",
            filters={
                "sort_by": "tenant_id",
                "sort_order": "asc",
            },
        )


def test_export_enforces_row_limit_before_workbook_creation(
    session,
):
    session.add_all(
        [
            SparePart(
                tenant_id="tenant-a",
                code="LIMIT-001",
                name="Limit One",
                unit="件",
            ),
            SparePart(
                tenant_id="tenant-a",
                code="LIMIT-002",
                name="Limit Two",
                unit="件",
            ),
        ]
    )
    session.commit()

    exporter = _exporter()
    exporter.max_rows = 1

    with pytest.raises(
        BusinessValidationError
    ) as raised:
        exporter.export(
            session,
            tenant_id="tenant-a",
            resource_key="spare-parts",
            filters={
                "sort_by": "code",
                "sort_order": "asc",
            },
        )

    assert (
        raised.value.code
        == "EXPORT_ROW_LIMIT_EXCEEDED"
    )

def test_configuration_export_excludes_cross_tenant_equipment_relation(
    session,
):
    from app.models.enums import ConfigurationStatus
    from app.models.equipment import (
        ConfigurationVersion,
        EquipmentModel,
    )

    other_tenant_model = EquipmentModel(
        tenant_id="tenant-b",
        code="MODEL-B",
        name="Tenant B Model",
    )
    session.add(other_tenant_model)
    session.flush()
    session.add(
        ConfigurationVersion(
            tenant_id="tenant-a",
            equipment_model_id=other_tenant_model.id,
            version_code="CFG-CROSS-TENANT",
            version_name="Cross Tenant",
            status=ConfigurationStatus.DRAFT,
        )
    )
    session.commit()

    content = _exporter().export(
        session,
        tenant_id="tenant-a",
        resource_key="configuration-versions",
        filters={},
    )
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
    )
    rows = _rows_as_dicts(
        workbook["构型版本"]
    )

    assert rows == [], (
        "CONFIGURATION_RELATED_TENANT_LEAK_GAP"
    )


def test_reliability_export_excludes_cross_tenant_spare_part_relation(
    session,
):
    from app.models.enums import (
        DataSourceType,
        ReliabilityModelType,
    )
    from app.models.reliability import (
        ReliabilityProfile,
    )

    other_tenant_part = SparePart(
        tenant_id="tenant-b",
        code="REL-PART-B",
        name="Tenant B Part",
        unit="件",
    )
    session.add(other_tenant_part)
    session.flush()
    session.add(
        ReliabilityProfile(
            tenant_id="tenant-a",
            profile_code="REL-CROSS-TENANT",
            spare_part_id=other_tenant_part.id,
            model_type=(
                ReliabilityModelType.EXPONENTIAL
            ),
            data_source_type=(
                DataSourceType.MANUAL_ESTIMATE
            ),
        )
    )
    session.commit()

    content = _exporter().export(
        session,
        tenant_id="tenant-a",
        resource_key="reliability-profiles",
        filters={},
    )
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
    )
    rows = _rows_as_dicts(
        workbook["可靠性参数"]
    )

    assert rows == [], (
        "RELIABILITY_RELATED_TENANT_LEAK_GAP"
    )


def test_inventory_export_excludes_each_cross_tenant_relation(
    session,
):
    from app.models import (
        InventoryBalance,
        Warehouse,
        WarehouseLocation,
    )

    tenant_a_part = SparePart(
        tenant_id="tenant-a",
        code="INV-PART-A",
        name="Tenant A Part",
        unit="件",
    )
    tenant_b_part = SparePart(
        tenant_id="tenant-b",
        code="INV-PART-B",
        name="Tenant B Part",
        unit="件",
    )
    tenant_a_warehouse = Warehouse(
        tenant_id="tenant-a",
        code="INV-WH-A",
        name="Tenant A Warehouse",
    )
    tenant_b_warehouse = Warehouse(
        tenant_id="tenant-b",
        code="INV-WH-B",
        name="Tenant B Warehouse",
    )
    session.add_all(
        [
            tenant_a_part,
            tenant_b_part,
            tenant_a_warehouse,
            tenant_b_warehouse,
        ]
    )
    session.flush()
    location_for_foreign_warehouse = WarehouseLocation(
        tenant_id="tenant-a",
        warehouse_id=tenant_b_warehouse.id,
        code="DEFAULT",
        name="Foreign warehouse location",
        location_type="DEFAULT",
    )
    location_for_local_warehouse = WarehouseLocation(
        tenant_id="tenant-a",
        warehouse_id=tenant_a_warehouse.id,
        code="DEFAULT",
        name="Local warehouse location",
        location_type="DEFAULT",
    )
    session.add_all(
        [location_for_foreign_warehouse, location_for_local_warehouse]
    )
    session.flush()
    session.add_all(
        [
            InventoryBalance(
                tenant_id="tenant-a",
                warehouse_id=tenant_b_warehouse.id,
                location_id=location_for_foreign_warehouse.id,
                spare_part_id=tenant_a_part.id,
                on_hand_quantity=1,
            ),
            InventoryBalance(
                tenant_id="tenant-a",
                warehouse_id=tenant_a_warehouse.id,
                location_id=location_for_local_warehouse.id,
                spare_part_id=tenant_b_part.id,
                on_hand_quantity=2,
            ),
        ]
    )
    session.commit()

    content = _exporter().export(
        session,
        tenant_id="tenant-a",
        resource_key="inventories",
        filters={},
    )
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
    )
    rows = _rows_as_dicts(workbook["库存"])

    assert rows == [], (
        "INVENTORY_RELATED_TENANT_LEAK_GAP"
    )


def test_inventory_export_aggregates_all_locations_and_lots_into_legacy_columns(
    session,
):
    from app.models import (
        InventoryBalance,
        InventoryLot,
        InventoryPolicy,
        Warehouse,
        WarehouseLocation,
    )

    warehouse = Warehouse(
        tenant_id="tenant-a",
        code="WH-AGG",
        name="Aggregate warehouse",
    )
    spare = SparePart(
        tenant_id="tenant-a",
        code="SP-AGG",
        name="Aggregate spare",
        unit="件",
    )
    session.add_all([warehouse, spare])
    session.flush()
    default = WarehouseLocation(
        tenant_id="tenant-a",
        warehouse_id=warehouse.id,
        code="DEFAULT",
        name="Default",
        location_type="DEFAULT",
    )
    shelf = WarehouseLocation(
        tenant_id="tenant-a",
        warehouse_id=warehouse.id,
        code="SHELF-1",
        name="Shelf 1",
        location_type="STORAGE",
    )
    lot = InventoryLot(
        tenant_id="tenant-a",
        spare_part_id=spare.id,
        lot_code="LOT-1",
    )
    session.add_all([default, shelf, lot])
    session.flush()
    session.add(
        InventoryPolicy(
            tenant_id="tenant-a",
            warehouse_id=warehouse.id,
            spare_part_id=spare.id,
            safety_stock=Decimal("4"),
            reorder_point=Decimal("6"),
            maximum_stock=Decimal("20"),
        )
    )
    session.add_all(
        [
            InventoryBalance(
                tenant_id="tenant-a",
                warehouse_id=warehouse.id,
                location_id=default.id,
                spare_part_id=spare.id,
                on_hand_quantity=Decimal("5"),
                reserved_quantity=Decimal("1"),
                in_transit_quantity=Decimal("2"),
            ),
            InventoryBalance(
                tenant_id="tenant-a",
                warehouse_id=warehouse.id,
                location_id=shelf.id,
                spare_part_id=spare.id,
                lot_id=lot.id,
                on_hand_quantity=Decimal("7"),
                damaged_quantity=Decimal("1"),
                quarantined_quantity=Decimal("0.5"),
                in_transit_quantity=Decimal("1"),
            ),
        ]
    )
    session.commit()

    content = _exporter().export(
        session,
        tenant_id="tenant-a",
        resource_key="inventories",
        filters={"sort_by": "on_hand_quantity"},
    )
    workbook = load_workbook(BytesIO(content), read_only=True)
    rows = _rows_as_dicts(workbook["库存"])

    assert rows == [
        {
            "库房编码": "WH-AGG",
            "器材编码": "SP-AGG",
            "现存数量": 12,
            "预留数量": 1,
            "损坏数量": 1,
            "隔离数量": 0.5,
            "在途数量": 3,
            "可用数量": 9.5,
            "安全库存": 4,
            "补货点": 6,
            "最大库存": 20,
            "盘点时间": None,
        }
    ]


def test_supplier_offer_export_excludes_each_cross_tenant_relation(
    session,
):
    from app.models.supplier import SupplierOffer

    tenant_a_part = SparePart(
        tenant_id="tenant-a",
        code="OFFER-PART-A",
        name="Tenant A Part",
        unit="件",
    )
    tenant_b_part = SparePart(
        tenant_id="tenant-b",
        code="OFFER-PART-B",
        name="Tenant B Part",
        unit="件",
    )
    tenant_a_supplier = Supplier(
        tenant_id="tenant-a",
        code="OFFER-SUP-A",
        name="Tenant A Supplier",
    )
    tenant_b_supplier = Supplier(
        tenant_id="tenant-b",
        code="OFFER-SUP-B",
        name="Tenant B Supplier",
    )
    session.add_all(
        [
            tenant_a_part,
            tenant_b_part,
            tenant_a_supplier,
            tenant_b_supplier,
        ]
    )
    session.flush()
    session.add_all(
        [
            SupplierOffer(
                tenant_id="tenant-a",
                offer_code="CROSS-SUPPLIER",
                supplier_id=tenant_b_supplier.id,
                spare_part_id=tenant_a_part.id,
                unit_price=1,
                lead_time_days=1,
            ),
            SupplierOffer(
                tenant_id="tenant-a",
                offer_code="CROSS-PART",
                supplier_id=tenant_a_supplier.id,
                spare_part_id=tenant_b_part.id,
                unit_price=2,
                lead_time_days=2,
            ),
        ]
    )
    session.commit()

    content = _exporter().export(
        session,
        tenant_id="tenant-a",
        resource_key="supplier-offers",
        filters={},
    )
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
    )
    rows = _rows_as_dicts(
        workbook["供应商报价"]
    )

    assert rows == [], (
        "SUPPLIER_OFFER_RELATED_TENANT_LEAK_GAP"
    )
