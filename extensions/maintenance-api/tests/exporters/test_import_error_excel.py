from __future__ import annotations

from io import BytesIO

from app.schemas.import_data import ImportIssue
from openpyxl import load_workbook


def _build_workbook() -> bytes:
    from app.exporters.import_error_excel import (
        build_import_error_workbook,
    )

    return build_import_error_workbook(
        errors=[
            ImportIssue(
                sheet="04_维修器材",
                row=2,
                field="code",
                code="REQUIRED",
                message="器材编码不能为空",
            ),
            ImportIssue(
                sheet="07_库房",
                row=5,
                field="name",
                code="INVALID_VALUE",
                message="库房名称无效",
            ),
        ],
        summaries=[
            {
                "name": "04_维修器材",
                "total_rows": 3,
                "valid_rows": 2,
                "invalid_rows": 1,
            },
            {
                "name": "07_库房",
                "total_rows": 5,
                "valid_rows": 4,
                "invalid_rows": 1,
            },
        ],
    )


def test_error_workbook_has_exact_sheets_and_headers():
    workbook = load_workbook(
        BytesIO(_build_workbook()),
        data_only=True,
    )

    assert workbook.sheetnames == [
        "导入错误",
        "导入摘要",
    ]
    assert [
        cell.value
        for cell in workbook["导入错误"][1]
    ] == [
        "工作表",
        "行号",
        "字段",
        "错误代码",
        "错误信息",
    ]
    assert [
        cell.value
        for cell in workbook["导入摘要"][1]
    ] == [
        "工作表",
        "总行数",
        "有效行数",
        "无效行数",
    ]


def test_error_workbook_preserves_issue_and_summary_values():
    workbook = load_workbook(
        BytesIO(_build_workbook()),
        data_only=True,
    )

    assert [
        cell.value
        for cell in workbook["导入错误"][2]
    ] == [
        "04_维修器材",
        2,
        "code",
        "REQUIRED",
        "器材编码不能为空",
    ]
    assert [
        cell.value
        for cell in workbook["导入错误"][3]
    ] == [
        "07_库房",
        5,
        "name",
        "INVALID_VALUE",
        "库房名称无效",
    ]
    assert [
        cell.value
        for cell in workbook["导入摘要"][2]
    ] == [
        "04_维修器材",
        3,
        2,
        1,
    ]
    assert [
        cell.value
        for cell in workbook["导入摘要"][3]
    ] == [
        "07_库房",
        5,
        4,
        1,
    ]
