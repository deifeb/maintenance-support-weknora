from __future__ import annotations

from collections.abc import Callable
from copy import deepcopy
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from app.models import (
    ConfigurationItem,
    ConfigurationVersion,
    EquipmentModel,
    InventoryBalance,
    InventoryLedgerEntry,
    InventoryPolicy,
    InventoryTargetReceipt,
    InventoryTargetReceiptStatus,
    InventoryTransaction,
    Part,
    SparePart,
    Warehouse,
    WarehouseLocation,
)
from app.models.enums import (
    ConfigurationStatus,
    CriticalityLevel,
    DemandExecutionMode,
    MissingParameterPolicy,
)
from app.schemas.inventory import InventoryQuantities
from app.security.actor import ActorContext, MaintenanceRole
from app.services.demand_calculation_service import DemandCalculationService
from app.services.inventory_target_adapter import InventoryTargetAdapter
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

_ADJUSTMENT_KEY = "plan05-4a-task7-adjust-1"
_TARGET_KEY = "plan05-4a-task7-import-target-1"
_QUANTITY_FIELDS = (
    "on_hand_quantity",
    "reserved_quantity",
    "damaged_quantity",
    "quarantined_quantity",
    "in_transit_quantity",
)
_EXPECTED_SUMMARY = {
    "on_hand_quantity": "12.0000",
    "reserved_quantity": "2.0000",
    "damaged_quantity": "1.0000",
    "quarantined_quantity": "1.0000",
    "in_transit_quantity": "3.0000",
    "available_quantity": "8.0000",
    "safety_stock": "2.0000",
    "reorder_point": "10.0000",
    "maximum_stock": "30.0000",
}
_EXPECTED_BEFORE = {
    "on_hand": "0.0000",
    "reserved": "0.0000",
    "damaged": "0.0000",
    "quarantined": "0.0000",
    "in_transit": "0.0000",
}
_EXPECTED_AFTER = {
    "on_hand": "12.0000",
    "reserved": "2.0000",
    "damaged": "1.0000",
    "quarantined": "1.0000",
    "in_transit": "3.0000",
}
_EXPORT_COLUMNS = {
    "现存数量": "12.0000",
    "预留数量": "2.0000",
    "损坏数量": "1.0000",
    "隔离数量": "1.0000",
    "在途数量": "3.0000",
    "可用数量": "8.0000",
    "安全库存": "2.0000",
    "补货点": "10.0000",
    "最大库存": "30.0000",
}


def _headers(
    internal_auth_headers: Callable[..., dict[str, str]],
    *,
    tenant_id: str,
    role: MaintenanceRole,
) -> dict[str, str]:
    return internal_auth_headers(
        tenant_id=tenant_id,
        user_id=f"{role.value}-{tenant_id}",
        role=role,
        request_id=f"plan05-4a-task7-{tenant_id}-{role.value}",
    )


def _scenario_version(configuration_version_id: int) -> SimpleNamespace:
    fleet = SimpleNamespace(
        group_code="FLEET-TASK7",
        configuration_version_id=configuration_version_id,
        initial_quantity=1,
        stage_usages=[],
        age_groups=[],
    )
    return SimpleNamespace(
        id=9701,
        stages=[],
        fleet_groups=[fleet],
        missing_parameter_policy=MissingParameterPolicy.FALLBACK,
        fallback_parameters_json={"failure_rate": "0.01"},
        default_service_level=Decimal("0.9"),
        default_initial_age_hours=Decimal("0"),
        execution_mode=DemandExecutionMode.ANALYTICAL,
        simulation_config_json={},
        formula_version="plan05-4a-task7",
        input_schema_version="1.0",
    )


def _seed_prerequisites(session: Session) -> SimpleNamespace:
    equipment = EquipmentModel(
        tenant_id="tenant-a",
        code="EQ-TASK7",
        name="Task 7 equipment",
    )
    configuration = ConfigurationVersion(
        tenant_id="tenant-a",
        equipment_model=equipment,
        version_code="TASK7-V1",
        version_name="Task 7 Version 1",
        status=ConfigurationStatus.PUBLISHED,
    )
    part = Part(
        tenant_id="tenant-a",
        code="PART-TASK7",
        name="Task 7 part",
    )
    spare = SparePart(
        tenant_id="tenant-a",
        code="SP-TASK7",
        name="Task 7 spare",
        unit="piece",
        is_repairable=False,
    )
    warehouse = Warehouse(
        tenant_id="tenant-a",
        code="WH-TASK7",
        name="Task 7 warehouse",
    )
    session.add_all((equipment, configuration, part, spare, warehouse))
    session.flush()
    session.add(
        ConfigurationItem(
            tenant_id="tenant-a",
            configuration_version_id=configuration.id,
            item_code="ITEM-TASK7",
            part_id=part.id,
            spare_part_id=spare.id,
            install_quantity=Decimal("1"),
            replacement_ratio=Decimal("1"),
            criticality_level=CriticalityLevel.HIGH,
        )
    )
    session.commit()
    return SimpleNamespace(
        configuration_id=configuration.id,
        warehouse_id=warehouse.id,
        warehouse_code=warehouse.code,
        spare_part_id=spare.id,
        spare_part_code=spare.code,
    )


def _create_inventory(
    client: TestClient,
    internal_auth_headers: Callable[..., dict[str, str]],
    seed: SimpleNamespace,
) -> dict:
    response = client.post(
        "/api/v1/master-data/inventories",
        headers=_headers(
            internal_auth_headers,
            tenant_id="tenant-a",
            role=MaintenanceRole.ADMIN,
        ),
        json={
            "warehouse_id": seed.warehouse_id,
            "spare_part_id": seed.spare_part_id,
            "safety_stock": "2.0000",
            "reorder_point": "10.0000",
            "maximum_stock": "30.0000",
            "notes": "Plan 05-4A Task 7",
        },
    )
    assert response.status_code == 201, response.text
    data = response.json()["data"]
    assert data["warehouse_id"] == seed.warehouse_id
    assert data["spare_part_id"] == seed.spare_part_id
    assert data["version"] == 1
    assert all(data[field] == "0.0000" for field in _QUANTITY_FIELDS)
    assert data["available_quantity"] == "0.0000"
    return data


def _adjust_inventory(
    client: TestClient,
    internal_auth_headers: Callable[..., dict[str, str]],
    balance_id: int,
    *,
    reason: str = "Plan 05-4A closure integration",
):
    return client.post(
        f"/api/v1/master-data/inventories/{balance_id}/adjust",
        headers={
            **_headers(
                internal_auth_headers,
                tenant_id="tenant-a",
                role=MaintenanceRole.ADMIN,
            ),
            "Idempotency-Key": _ADJUSTMENT_KEY,
        },
        json={
            "expected_version": 1,
            "on_hand_delta": "12.0000",
            "reserved_delta": "2.0000",
            "damaged_delta": "1.0000",
            "quarantined_delta": "1.0000",
            "in_transit_delta": "3.0000",
            "reason": reason,
        },
    )


def _assert_summary(summary: dict) -> None:
    for key, expected in _EXPECTED_SUMMARY.items():
        assert summary[key] == expected


def _record_import_target_receipt(
    session: Session,
    actor_context: Callable[..., ActorContext],
    seed: SimpleNamespace,
):
    actor = actor_context(
        tenant_id="tenant-a",
        user_id="admin-tenant-a",
        role=MaintenanceRole.ADMIN,
        request_id="plan05-4a-task7-target",
        token_id="plan05-4a-task7-target-token",
    )
    source_payload = {
        "resource": "inventories",
        "warehouse_id": seed.warehouse_id,
        "spare_part_id": seed.spare_part_id,
        "target": deepcopy(_EXPECTED_SUMMARY),
    }
    result = InventoryTargetAdapter().apply_target(
        session,
        actor,
        warehouse_id=seed.warehouse_id,
        spare_part_id=seed.spare_part_id,
        quantities=InventoryQuantities(
            on_hand_quantity=Decimal("12.0000"),
            reserved_quantity=Decimal("2.0000"),
            damaged_quantity=Decimal("1.0000"),
            quarantined_quantity=Decimal("1.0000"),
            in_transit_quantity=Decimal("3.0000"),
            safety_stock=Decimal("2.0000"),
            reorder_point=Decimal("10.0000"),
            maximum_stock=Decimal("30.0000"),
        ),
        notes="Plan 05-4A Task 7",
        idempotency_key=_TARGET_KEY,
        source_payload=source_payload,
        reason="Plan 05-4A closure import compatibility",
    )
    session.commit()
    return result, source_payload, actor


def _workbook_rows(content: bytes) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(BytesIO(content), data_only=False)
    worksheet = workbook["库存"]
    headers = [cell.value for cell in worksheet[1]]
    rows = [
        dict(zip(headers, values, strict=True))
        for values in worksheet.iter_rows(min_row=2, values_only=True)
    ]
    return headers, rows


def _tenant_fact_counts(session: Session, tenant_id: str) -> dict[str, int]:
    models = {
        "balance": InventoryBalance,
        "policy": InventoryPolicy,
        "transaction": InventoryTransaction,
        "entry": InventoryLedgerEntry,
        "receipt": InventoryTargetReceipt,
    }
    return {
        name: session.scalar(
            select(func.count()).select_from(model).where(
                model.tenant_id == tenant_id
            )
        )
        or 0
        for name, model in models.items()
    }


def test_authoritative_inventory_fact_flows_through_all_05_4a_consumers(
    client: TestClient,
    session: Session,
    actor_context: Callable[..., ActorContext],
    internal_auth_headers: Callable[..., dict[str, str]],
) -> None:
    seed = _seed_prerequisites(session)
    created = _create_inventory(client, internal_auth_headers, seed)
    balance_id = created["id"]

    session.expire_all()
    balance = session.get(InventoryBalance, balance_id)
    assert balance is not None
    location = session.get(WarehouseLocation, balance.location_id)
    assert location is not None
    assert location.code == "DEFAULT"

    adjusted = _adjust_inventory(client, internal_auth_headers, balance_id)
    assert adjusted.status_code == 200, adjusted.text
    adjusted_data = adjusted.json()["data"]
    _assert_summary(adjusted_data["summary"])

    fetched = client.get(
        f"/api/v1/master-data/inventories/{balance_id}",
        headers=_headers(
            internal_auth_headers,
            tenant_id="tenant-a",
            role=MaintenanceRole.VIEWER,
        ),
    )
    assert fetched.status_code == 200, fetched.text
    _assert_summary(fetched.json()["data"])

    dashboard = client.get(
        "/api/v1/dashboard/summary",
        headers=_headers(
            internal_auth_headers,
            tenant_id="tenant-a",
            role=MaintenanceRole.VIEWER,
        ),
    )
    assert dashboard.status_code == 200, dashboard.text
    metrics = {
        metric["key"]: metric["value"]
        for metric in dashboard.json()["data"]["metrics"]
    }
    assert metrics["inventory_risk_count"] == 1

    session.expire_all()
    actor = actor_context(
        tenant_id="tenant-a",
        role=MaintenanceRole.VIEWER,
    )
    snapshot, _warnings = DemandCalculationService()._snapshot_from_version(
        session,
        actor,
        _scenario_version(seed.configuration_id),
    )
    snapshot_item = next(
        item
        for item in snapshot["items"]
        if item["spare_part_id"] == seed.spare_part_id
    )
    assert snapshot_item["inventory"] == {
        "on_hand_quantity": "12.0000",
        "available_quantity": "8.0000",
        "in_transit_quantity": "3.0000",
        "safety_stock": "2.0000",
    }

    exported = client.get(
        "/api/v1/master-data/exports/inventories",
        params={
            "warehouse_id": seed.warehouse_id,
            "spare_part_id": seed.spare_part_id,
            "sort_by": "available_quantity",
            "sort_order": "asc",
        },
        headers=_headers(
            internal_auth_headers,
            tenant_id="tenant-a",
            role=MaintenanceRole.VIEWER,
        ),
    )
    assert exported.status_code == 200, exported.text
    _headers_row, rows = _workbook_rows(exported.content)
    assert len(rows) == 1
    assert rows[0]["库房编码"] == seed.warehouse_code
    assert rows[0]["器材编码"] == seed.spare_part_code
    for column, expected in _EXPORT_COLUMNS.items():
        assert rows[0][column] == expected
        assert isinstance(rows[0][column], str)

    receipt_result, _source_payload, _actor = _record_import_target_receipt(
        session,
        actor_context,
        seed,
    )
    assert receipt_result.created_identity is False
    assert receipt_result.operation_type is None
    assert receipt_result.transaction_id is None
    assert receipt_result.replayed is False

    session.expire_all()
    transaction = session.scalar(
        select(InventoryTransaction).where(
            InventoryTransaction.tenant_id == "tenant-a",
            InventoryTransaction.idempotency_key == _ADJUSTMENT_KEY,
        )
    )
    assert transaction is not None
    assert transaction.status == "COMPLETED"
    entry = session.scalar(
        select(InventoryLedgerEntry).where(
            InventoryLedgerEntry.tenant_id == "tenant-a",
            InventoryLedgerEntry.transaction_id == transaction.id,
        )
    )
    assert entry is not None
    receipt = session.scalar(
        select(InventoryTargetReceipt).where(
            InventoryTargetReceipt.tenant_id == "tenant-a",
            InventoryTargetReceipt.idempotency_key == _TARGET_KEY,
        )
    )
    assert receipt is not None
    assert receipt.status is InventoryTargetReceiptStatus.COMPLETED
    assert receipt.result_json == {
        "created_identity": False,
        "operation_type": None,
        "transaction_id": None,
    }
    assert session.scalar(
        select(func.count()).select_from(InventoryTransaction).where(
            InventoryTransaction.tenant_id == "tenant-a",
            InventoryTransaction.idempotency_key == _ADJUSTMENT_KEY,
        )
    ) == 1
    assert session.scalar(
        select(func.count()).select_from(InventoryLedgerEntry).where(
            InventoryLedgerEntry.tenant_id == "tenant-a",
            InventoryLedgerEntry.transaction_id == transaction.id,
        )
    ) == 1
    assert session.scalar(
        select(func.count()).select_from(InventoryTargetReceipt).where(
            InventoryTargetReceipt.tenant_id == "tenant-a",
            InventoryTargetReceipt.idempotency_key == _TARGET_KEY,
        )
    ) == 1
    assert (
        entry.on_hand_delta,
        entry.reserved_delta,
        entry.damaged_delta,
        entry.quarantined_delta,
        entry.in_transit_delta,
    ) == (
        Decimal("12.0000"),
        Decimal("2.0000"),
        Decimal("1.0000"),
        Decimal("1.0000"),
        Decimal("3.0000"),
    )
    assert entry.state_before_json == _EXPECTED_BEFORE
    assert entry.state_after_json == _EXPECTED_AFTER

    balance = session.get(InventoryBalance, balance_id)
    policy = session.scalar(
        select(InventoryPolicy).where(
            InventoryPolicy.tenant_id == "tenant-a",
            InventoryPolicy.warehouse_id == seed.warehouse_id,
            InventoryPolicy.spare_part_id == seed.spare_part_id,
        )
    )
    warehouse = session.get(Warehouse, seed.warehouse_id)
    spare = session.get(SparePart, seed.spare_part_id)
    assert all(
        item is not None and item.tenant_id == "tenant-a"
        for item in (
            transaction,
            entry,
            receipt,
            balance,
            policy,
            warehouse,
            spare,
        )
    )


def test_authoritative_inventory_adjustment_replay_is_side_effect_free(
    client: TestClient,
    session: Session,
    actor_context: Callable[..., ActorContext],
    internal_auth_headers: Callable[..., dict[str, str]],
) -> None:
    seed = _seed_prerequisites(session)
    created = _create_inventory(client, internal_auth_headers, seed)
    balance_id = created["id"]

    first = _adjust_inventory(client, internal_auth_headers, balance_id)
    assert first.status_code == 200, first.text
    receipt_first, source_payload, actor = _record_import_target_receipt(
        session,
        actor_context,
        seed,
    )
    assert receipt_first.replayed is False

    session.expire_all()
    transaction = session.scalar(
        select(InventoryTransaction).where(
            InventoryTransaction.tenant_id == "tenant-a",
            InventoryTransaction.idempotency_key == _ADJUSTMENT_KEY,
        )
    )
    assert transaction is not None
    recorded = {
        "response": first.json(),
        "counts": {
            "transactions": session.scalar(
                select(func.count()).select_from(InventoryTransaction)
            ),
            "entries": session.scalar(
                select(func.count()).select_from(InventoryLedgerEntry)
            ),
            "receipts": session.scalar(
                select(func.count()).select_from(InventoryTargetReceipt)
            ),
        },
        "snapshot": deepcopy(transaction.response_snapshot_json),
    }

    replay = _adjust_inventory(client, internal_auth_headers, balance_id)
    assert replay.status_code == 200, replay.text
    assert replay.json() == recorded["response"]

    target_replay = InventoryTargetAdapter().apply_target(
        session,
        actor,
        warehouse_id=seed.warehouse_id,
        spare_part_id=seed.spare_part_id,
        quantities=InventoryQuantities(
            on_hand_quantity=Decimal("12.0000"),
            reserved_quantity=Decimal("2.0000"),
            damaged_quantity=Decimal("1.0000"),
            quarantined_quantity=Decimal("1.0000"),
            in_transit_quantity=Decimal("3.0000"),
            safety_stock=Decimal("2.0000"),
            reorder_point=Decimal("10.0000"),
            maximum_stock=Decimal("30.0000"),
        ),
        notes="Plan 05-4A Task 7",
        idempotency_key=_TARGET_KEY,
        source_payload=source_payload,
        reason="Plan 05-4A closure import compatibility",
    )
    session.commit()
    assert target_replay.replayed is True

    different = _adjust_inventory(
        client,
        internal_auth_headers,
        balance_id,
        reason="Plan 05-4A different closure request",
    )
    assert different.status_code == 409
    assert different.json()["error"]["code"] == "IDEMPOTENCY_KEY_REUSED"

    session.expire_all()
    transaction = session.get(InventoryTransaction, transaction.id)
    assert transaction is not None
    assert {
        "transactions": session.scalar(
            select(func.count()).select_from(InventoryTransaction)
        ),
        "entries": session.scalar(
            select(func.count()).select_from(InventoryLedgerEntry)
        ),
        "receipts": session.scalar(
            select(func.count()).select_from(InventoryTargetReceipt)
        ),
    } == recorded["counts"]
    assert transaction.response_snapshot_json == recorded["snapshot"]


def test_authoritative_inventory_fact_is_hidden_from_foreign_tenant(
    client: TestClient,
    session: Session,
    internal_auth_headers: Callable[..., dict[str, str]],
) -> None:
    seed = _seed_prerequisites(session)
    created = _create_inventory(client, internal_auth_headers, seed)
    balance_id = created["id"]
    adjusted = _adjust_inventory(client, internal_auth_headers, balance_id)
    assert adjusted.status_code == 200, adjusted.text

    session.expire_all()
    before = _tenant_fact_counts(session, "tenant-b")
    foreign_headers = _headers(
        internal_auth_headers,
        tenant_id="tenant-b",
        role=MaintenanceRole.VIEWER,
    )

    fetched = client.get(
        f"/api/v1/master-data/inventories/{balance_id}",
        headers=foreign_headers,
    )
    assert fetched.status_code == 404

    exported = client.get(
        "/api/v1/master-data/exports/inventories",
        params={
            "warehouse_id": seed.warehouse_id,
            "spare_part_id": seed.spare_part_id,
        },
        headers=foreign_headers,
    )
    assert exported.status_code == 200, exported.text
    headers_row, rows = _workbook_rows(exported.content)
    assert rows == []
    assert seed.warehouse_code not in headers_row
    assert seed.spare_part_code not in headers_row

    dashboard = client.get(
        "/api/v1/dashboard/summary",
        headers=foreign_headers,
    )
    assert dashboard.status_code == 200, dashboard.text
    metrics = {
        metric["key"]: metric["value"]
        for metric in dashboard.json()["data"]["metrics"]
    }
    assert metrics["inventory_risk_count"] == 0

    session.expire_all()
    assert _tenant_fact_counts(session, "tenant-b") == before
