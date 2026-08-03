from __future__ import annotations

from io import BytesIO

from app.main import create_app
from app.models.catalog import SparePart
from app.security.actor import MaintenanceRole
from fastapi.routing import (
    APIRoute,
    iter_route_contexts,
)
from openpyxl import load_workbook


def _rows_as_dicts(worksheet) -> list[dict[str, object]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    return [
        dict(zip(headers, row, strict=True))
        for row in rows
    ]


def test_viewer_export_route_is_tenant_filtered_and_sets_headers(
    client,
    session,
    internal_auth_headers,
):
    session.add_all(
        [
            SparePart(
                tenant_id="tenant-a",
                code="SP-SHARED",
                name="Tenant A Part",
                unit="件",
                is_critical=True,
            ),
            SparePart(
                tenant_id="tenant-b",
                code="SP-SHARED",
                name="Tenant B Part",
                unit="件",
                is_critical=True,
            ),
        ]
    )
    session.commit()

    response = client.get(
        (
            "/api/v1/master-data/exports/"
            "spare-parts"
            "?keyword=SP-SHARED"
            "&is_critical=true"
            "&sort_by=code"
            "&sort_order=asc"
        ),
        headers=internal_auth_headers(
            tenant_id="tenant-a",
            user_id="viewer-a",
            role=MaintenanceRole.VIEWER,
        ),
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    disposition = response.headers[
        "content-disposition"
    ]
    assert disposition.startswith(
        "attachment; filename*=UTF-8''"
    )
    assert disposition.endswith(".xlsx")
    assert response.headers["cache-control"] == "no-store"

    workbook = load_workbook(
        BytesIO(response.content),
        read_only=True,
    )
    rows = _rows_as_dicts(
        workbook["维修器材"]
    )

    assert len(rows) == 1
    assert rows[0]["器材编码"] == "SP-SHARED"
    assert rows[0]["器材名称"] == "Tenant A Part"


def test_export_route_rejects_unauthenticated_request(
    client,
):
    response = client.get(
        (
            "/api/v1/master-data/exports/"
            "spare-parts"
        )
    )

    assert response.status_code == 401


def test_export_route_requires_viewer_and_declares_no_tenant_parameter(
):
    app = create_app()
    matches = []

    for route in iter_route_contexts(
        app.routes
    ):
        if not isinstance(
            route.original_route,
            APIRoute,
        ):
            continue
        if "GET" not in (route.methods or set()):
            continue
        if not route.path.startswith(
            "/api/v1/master-data/exports/"
        ):
            continue
        matches.append(route)

    assert len(matches) == 1

    route = matches[0]
    dependency_names = {
        dependency.call.__name__
        for dependency
        in route.dependant.dependencies
        if dependency.call is not None
        and hasattr(
            dependency.call,
            "__name__",
        )
    }
    query_names = {
        parameter.name
        for parameter
        in route.dependant.query_params
    }

    assert "require_viewer" in dependency_names
    assert "require_contributor" not in dependency_names
    assert "require_admin" not in dependency_names
    assert "tenant_id" not in query_names


def test_export_route_rejects_unknown_sort_field(
    client,
    internal_auth_headers,
):
    response = client.get(
        (
            "/api/v1/master-data/exports/"
            "spare-parts"
            "?sort_by=tenant_id"
        ),
        headers=internal_auth_headers(
            tenant_id="tenant-a",
            user_id="viewer-a",
            role=MaintenanceRole.VIEWER,
        ),
    )

    assert response.status_code == 422

def test_export_route_declares_exact_approved_query_contract():
    app = create_app()
    matches = []

    for route in iter_route_contexts(app.routes):
        if not isinstance(
            route.original_route,
            APIRoute,
        ):
            continue
        if "GET" not in (route.methods or set()):
            continue
        if route.path != (
            "/api/v1/master-data/exports/"
            "{resource_key}"
        ):
            continue
        matches.append(route)

    assert len(matches) == 1

    query_names = {
        parameter.name
        for parameter
        in matches[0].dependant.query_params
    }
    assert query_names == {
        "keyword",
        "include_inactive",
        "is_critical",
        "is_repairable",
        "spare_part_id",
        "warehouse_id",
        "supplier_id",
        "sort_by",
        "sort_order",
    }, "APPROVED_EXPORT_QUERY_CONTRACT_GAP"


def test_export_route_include_inactive_false_excludes_inactive_rows(
    client,
    session,
    internal_auth_headers,
):
    session.add_all(
        [
            SparePart(
                tenant_id="tenant-a",
                code="ACTIVE-ONLY",
                name="Active",
                unit="件",
                is_active=True,
            ),
            SparePart(
                tenant_id="tenant-a",
                code="INACTIVE-HIDDEN",
                name="Inactive",
                unit="件",
                is_active=False,
            ),
        ]
    )
    session.commit()

    response = client.get(
        (
            "/api/v1/master-data/exports/"
            "spare-parts"
            "?include_inactive=false"
            "&sort_by=code"
            "&sort_order=asc"
        ),
        headers=internal_auth_headers(
            tenant_id="tenant-a",
            user_id="viewer-a",
            role=MaintenanceRole.VIEWER,
        ),
    )

    assert response.status_code == 200
    workbook = load_workbook(
        BytesIO(response.content),
        read_only=True,
    )
    rows = _rows_as_dicts(
        workbook["维修器材"]
    )

    assert [
        row["器材编码"]
        for row in rows
    ] == ["ACTIVE-ONLY"], (
        "INCLUDE_INACTIVE_FILTER_GAP"
    )


def test_export_route_filters_reliability_profiles_by_spare_part_id(
    client,
    session,
    internal_auth_headers,
):
    from app.models.enums import (
        DataSourceType,
        ReliabilityModelType,
    )
    from app.models.reliability import (
        ReliabilityProfile,
    )

    target = SparePart(
        tenant_id="tenant-a",
        code="REL-TARGET",
        name="Target",
        unit="件",
    )
    other = SparePart(
        tenant_id="tenant-a",
        code="REL-OTHER",
        name="Other",
        unit="件",
    )
    session.add_all([target, other])
    session.flush()
    session.add_all(
        [
            ReliabilityProfile(
                tenant_id="tenant-a",
                profile_code="REL-001",
                spare_part_id=target.id,
                model_type=(
                    ReliabilityModelType.EXPONENTIAL
                ),
                data_source_type=(
                    DataSourceType.MANUAL_ESTIMATE
                ),
            ),
            ReliabilityProfile(
                tenant_id="tenant-a",
                profile_code="REL-002",
                spare_part_id=other.id,
                model_type=(
                    ReliabilityModelType.EXPONENTIAL
                ),
                data_source_type=(
                    DataSourceType.MANUAL_ESTIMATE
                ),
            ),
        ]
    )
    session.commit()

    response = client.get(
        (
            "/api/v1/master-data/exports/"
            "reliability-profiles"
            f"?spare_part_id={target.id}"
            "&sort_by=profile_code"
            "&sort_order=asc"
        ),
        headers=internal_auth_headers(
            tenant_id="tenant-a",
            user_id="viewer-a",
            role=MaintenanceRole.VIEWER,
        ),
    )

    assert response.status_code == 200
    workbook = load_workbook(
        BytesIO(response.content),
        read_only=True,
    )
    rows = _rows_as_dicts(
        workbook["可靠性参数"]
    )

    assert [
        row["参数档案编码"]
        for row in rows
    ] == ["REL-001"], (
        "SPARE_PART_ID_EXPORT_FILTER_GAP"
    )


def test_export_route_filters_inventories_by_spare_part_and_warehouse(
    client,
    session,
    internal_auth_headers,
):
    from app.models.inventory import (
        Warehouse,
        WarehouseInventory,
    )

    target_part = SparePart(
        tenant_id="tenant-a",
        code="INV-PART-TARGET",
        name="Target Part",
        unit="件",
    )
    other_part = SparePart(
        tenant_id="tenant-a",
        code="INV-PART-OTHER",
        name="Other Part",
        unit="件",
    )
    target_warehouse = Warehouse(
        tenant_id="tenant-a",
        code="WH-TARGET",
        name="Target Warehouse",
    )
    other_warehouse = Warehouse(
        tenant_id="tenant-a",
        code="WH-OTHER",
        name="Other Warehouse",
    )
    session.add_all(
        [
            target_part,
            other_part,
            target_warehouse,
            other_warehouse,
        ]
    )
    session.flush()
    session.add_all(
        [
            WarehouseInventory(
                tenant_id="tenant-a",
                warehouse_id=target_warehouse.id,
                spare_part_id=target_part.id,
                on_hand_quantity=1,
            ),
            WarehouseInventory(
                tenant_id="tenant-a",
                warehouse_id=other_warehouse.id,
                spare_part_id=other_part.id,
                on_hand_quantity=2,
            ),
        ]
    )
    session.commit()

    response = client.get(
        (
            "/api/v1/master-data/exports/"
            "inventories"
            f"?spare_part_id={target_part.id}"
            f"&warehouse_id={target_warehouse.id}"
            "&sort_by=on_hand_quantity"
            "&sort_order=asc"
        ),
        headers=internal_auth_headers(
            tenant_id="tenant-a",
            user_id="viewer-a",
            role=MaintenanceRole.VIEWER,
        ),
    )

    assert response.status_code == 200
    workbook = load_workbook(
        BytesIO(response.content),
        read_only=True,
    )
    rows = _rows_as_dicts(workbook["库存"])

    assert [
        (
            row["库房编码"],
            row["器材编码"],
        )
        for row in rows
    ] == [
        (
            "WH-TARGET",
            "INV-PART-TARGET",
        )
    ], "INVENTORY_ID_EXPORT_FILTER_GAP"


def test_export_route_filters_supplier_offers_by_supplier_and_spare_part(
    client,
    session,
    internal_auth_headers,
):
    from app.models.supplier import (
        Supplier,
        SupplierOffer,
    )

    target_part = SparePart(
        tenant_id="tenant-a",
        code="OFFER-PART-TARGET",
        name="Target Part",
        unit="件",
    )
    other_part = SparePart(
        tenant_id="tenant-a",
        code="OFFER-PART-OTHER",
        name="Other Part",
        unit="件",
    )
    target_supplier = Supplier(
        tenant_id="tenant-a",
        code="SUP-TARGET",
        name="Target Supplier",
    )
    other_supplier = Supplier(
        tenant_id="tenant-a",
        code="SUP-OTHER",
        name="Other Supplier",
    )
    session.add_all(
        [
            target_part,
            other_part,
            target_supplier,
            other_supplier,
        ]
    )
    session.flush()
    session.add_all(
        [
            SupplierOffer(
                tenant_id="tenant-a",
                offer_code="OFFER-001",
                supplier_id=target_supplier.id,
                spare_part_id=target_part.id,
                unit_price=1,
                lead_time_days=1,
            ),
            SupplierOffer(
                tenant_id="tenant-a",
                offer_code="OFFER-002",
                supplier_id=other_supplier.id,
                spare_part_id=other_part.id,
                unit_price=2,
                lead_time_days=2,
            ),
        ]
    )
    session.commit()

    response = client.get(
        (
            "/api/v1/master-data/exports/"
            "supplier-offers"
            f"?spare_part_id={target_part.id}"
            f"&supplier_id={target_supplier.id}"
            "&sort_by=offer_code"
            "&sort_order=asc"
        ),
        headers=internal_auth_headers(
            tenant_id="tenant-a",
            user_id="viewer-a",
            role=MaintenanceRole.VIEWER,
        ),
    )

    assert response.status_code == 200
    workbook = load_workbook(
        BytesIO(response.content),
        read_only=True,
    )
    rows = _rows_as_dicts(
        workbook["供应商报价"]
    )

    assert [
        row["报价编码"]
        for row in rows
    ] == ["OFFER-001"], (
        "SUPPLIER_OFFER_ID_EXPORT_FILTER_GAP"
    )


def test_export_route_uses_master_data_export_max_rows_setting(
    client,
    session,
    internal_auth_headers,
    monkeypatch,
):
    import importlib

    from app.api.v1.master_data import (
        exports as exports_module,
    )
    from app.core.config import get_settings

    session.add_all(
        [
            SparePart(
                tenant_id="tenant-a",
                code=f"CONFIG-LIMIT-{index:04d}",
                name=f"Configured Limit {index}",
                unit="件",
            )
            for index in range(1001)
        ]
    )
    session.commit()

    with monkeypatch.context() as patch:
        patch.setenv(
            "MASTER_DATA_EXPORT_MAX_ROWS",
            "1000",
        )
        get_settings.cache_clear()
        importlib.reload(exports_module)

        response = client.get(
            (
                "/api/v1/master-data/exports/"
                "spare-parts"
                "?sort_by=code"
                "&sort_order=asc"
            ),
            headers=internal_auth_headers(
                tenant_id="tenant-a",
                user_id="viewer-a",
                role=MaintenanceRole.VIEWER,
            ),
        )

    get_settings.cache_clear()
    importlib.reload(exports_module)

    assert response.status_code == 422, (
        "CONFIGURED_EXPORT_ROW_LIMIT_GAP"
    )
    assert (
        response.json()["error"]["code"]
        == "EXPORT_ROW_LIMIT_EXCEEDED"
    ), "CONFIGURED_EXPORT_ROW_LIMIT_CODE_GAP"
