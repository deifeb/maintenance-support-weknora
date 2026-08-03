from __future__ import annotations

import hashlib
import importlib
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from app.importers.parser import WorkbookParser
from app.importers.template import create_template_bytes
from app.models.import_task import ImportTaskStatus
from app.repositories.import_task_repository import (
    ImportTaskRepository,
)
from app.services.import_task_service import (
    ImportTaskFileStore,
    ImportTaskService,
)
from openpyxl import load_workbook


def test_settings_expose_exact_master_data_transfer_limits():
    from app.core.config import SERVICE_ROOT, Settings

    settings = Settings()

    assert settings.master_data_import_dir == (
        SERVICE_ROOT / "exports" / "master-data-imports"
    )
    assert settings.master_data_import_task_ttl_seconds == 1800
    assert settings.master_data_import_worker_count == 1
    assert settings.master_data_import_max_pending_tasks == 10
    assert settings.master_data_export_max_rows == 100_000


def test_env_example_exposes_limits_without_browser_file_path():
    service_root = Path(__file__).resolve().parents[2]
    env_text = (service_root / ".env.example").read_text(
        encoding="utf-8"
    )

    required_lines = {
        "MASTER_DATA_IMPORT_TASK_TTL_SECONDS=1800",
        "MASTER_DATA_IMPORT_WORKER_COUNT=1",
        "MASTER_DATA_IMPORT_MAX_PENDING_TASKS=10",
        "MASTER_DATA_EXPORT_MAX_ROWS=100000",
    }

    assert required_lines <= set(env_text.splitlines())
    assert "MASTER_DATA_IMPORT_DIR" not in env_text


def test_import_task_model_matches_approved_plan_contract():
    from enum import StrEnum

    module = importlib.import_module("app.models.import_task")
    model = getattr(module, "MasterDataImportTask", None)

    assert model is not None
    assert issubclass(ImportTaskStatus, StrEnum)

    assert set(model.__table__.columns.keys()) == {
        "id",
        "tenant_id",
        "created_by_user_id",
        "created_by_request_id",
        "original_filename",
        "file_path",
        "file_sha256",
        "template_version",
        "status",
        "mapping_json",
        "sheet_summary_json",
        "preview_json",
        "errors_json",
        "warnings_json",
        "result_json",
        "error_code",
        "error_message",
        "error_workbook_path",
        "expires_at",
        "started_at",
        "finished_at",
        "created_at",
        "updated_at",
        "version",
    }

    index_columns = {
        tuple(column.name for column in index.columns)
        for index in model.__table__.indexes
    }
    assert (
        "tenant_id",
        "created_by_user_id",
        "status",
    ) in index_columns
    assert ("expires_at",) in index_columns


def test_import_task_types_are_exported_from_public_packages():
    import app.models as models
    import app.repositories as repositories

    assert hasattr(models, "MasterDataImportTask")
    assert hasattr(models, "ImportTaskStatus")
    assert hasattr(repositories, "ImportTaskRepository")
    assert hasattr(repositories, "import_task_repository")


def test_file_store_supports_source_error_hash_and_cleanup_contract(
    tmp_path,
):
    store = ImportTaskFileStore(root=tmp_path)
    task_id = "12345678-1234-5678-1234-567812345678"
    source = b"source workbook bytes"
    errors = b"error workbook bytes"

    source_path = store.write_source(
        task_id=task_id,
        original_filename="../../escape.xlsx",
        content=source,
    )
    error_path = store.write_error_workbook(
        task_id=task_id,
        content=errors,
    )

    assert source_path == f"{task_id}/source.xlsx"
    assert error_path == f"{task_id}/errors.xlsx"
    assert store.read_source(source_path) == source
    assert store.read_error_workbook(error_path) == errors
    assert store.sha256(source) == hashlib.sha256(source).hexdigest()

    store.delete_task_files(task_id)

    assert not (tmp_path / task_id).exists()


def test_inspect_workbook_returns_plan_header_mapping():
    inspection = importlib.import_module(
        "app.importers.inspection"
    )
    result = inspection.inspect_workbook(create_template_bytes())

    assert result["template_version"] == "PLAN05-2-TASK09-V1"

    spare_sheet = next(
        sheet
        for sheet in result["sheets"]
        if sheet["name"] == "04_维修器材"
    )

    assert spare_sheet["source_headers"][:4] == [
        "操作",
        "器材编码",
        "器材名称",
        "规格型号",
    ]
    assert spare_sheet["suggested_mapping"]["操作"] == "operation"
    assert spare_sheet["suggested_mapping"]["器材编码"] == "code"
    assert spare_sheet["suggested_mapping"]["器材名称"] == "name"
    assert spare_sheet["suggested_mapping"]["单位"] == "unit"
    assert spare_sheet["required_fields"] == [
        "operation",
        "code",
        "name",
    ]


def test_parser_accepts_source_header_to_canonical_mapping():
    workbook = load_workbook(BytesIO(create_template_bytes()))
    sheet = workbook["04_维修器材"]

    sheet.cell(row=1, column=1, value="动作")
    sheet.cell(row=1, column=2, value="自定义编码")
    sheet.cell(row=1, column=3, value="自定义名称")
    sheet.cell(row=2, column=1, value="CREATE")
    sheet.cell(row=2, column=2, value="SP-MAPPED")
    sheet.cell(row=2, column=3, value="映射器材")

    output = BytesIO()
    workbook.save(output)

    parser = WorkbookParser(
        max_size_mb=10,
        max_rows_per_sheet=10_000,
    )
    parsed, errors = parser.parse(
        output.getvalue(),
        "mapped.xlsx",
        mapping={
            "04_维修器材": {
                "动作": "operation",
                "自定义编码": "code",
                "自定义名称": "name",
            }
        },
    )

    assert errors == []
    assert parsed["04_维修器材"] == [
        {
            "operation": "CREATE",
            "code": "SP-MAPPED",
            "name": "映射器材",
            "_row": 2,
        }
    ]


def test_upload_commits_inspected_task_metadata(
    session,
    actor_admin,
):
    with TemporaryDirectory() as directory:
        service = ImportTaskService(
            repository=ImportTaskRepository(),
            file_store=ImportTaskFileStore(
                root=Path(directory)
            ),
            task_ttl_seconds=1800,
            max_size_mb=10,
        )

        task = service.upload(
            session,
            actor=actor_admin,
            content=create_template_bytes(),
            filename="master-data.xlsx",
        )

        assert not session.in_transaction()
        assert task.status is ImportTaskStatus.UPLOADED
        assert task.template_version == "PLAN05-2-TASK09-V1"
        assert task.sheet_summary_json["04_维修器材"] == 0
        assert task.mapping_json is None
        assert task.preview_json is None
        assert task.errors_json is None
        assert task.warnings_json is None
        assert task.result_json is None
