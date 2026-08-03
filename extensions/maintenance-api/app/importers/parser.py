import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.importers.template import SHEET_SPECS, header_map
from app.models.enums import ImportOperation
from app.schemas.import_data import ImportIssue

TRUE_VALUES = {"1", "TRUE", "YES", "Y", "是", "启用"}
FALSE_VALUES = {"0", "FALSE", "NO", "N", "否", "停用"}


def normalize_code(value: Any) -> str:
    return str(value).strip().upper()


def parse_bool(
    value: Any,
    default: bool | None = None,
) -> bool | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().upper()
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    raise ValueError("must be a boolean value")


def parse_decimal(
    value: Any,
    default: Decimal | None = None,
) -> Decimal | None:
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(
            "must be a decimal number"
        ) from exc


def parse_int(
    value: Any,
    default: int | None = None,
) -> int | None:
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("must be an integer") from exc


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value).strip())


def parse_json(value: Any) -> dict[str, Any] | None:
    if value is None or value == "":
        return None
    if isinstance(value, dict):
        return value
    parsed = json.loads(str(value))
    if not isinstance(parsed, dict):
        raise ValueError("must be a JSON object")
    return parsed


class WorkbookParser:
    def __init__(
        self,
        *,
        max_size_mb: int,
        max_rows_per_sheet: int,
    ) -> None:
        self.max_size_bytes = max_size_mb * 1024 * 1024
        self.max_rows_per_sheet = max_rows_per_sheet

    def parse(
        self,
        content: bytes,
        filename: str,
        mapping: dict[str, dict[str, str]] | None = None,
    ) -> tuple[
        dict[str, list[dict[str, Any]]],
        list[ImportIssue],
    ]:
        errors: list[ImportIssue] = []
        if not filename.lower().endswith(".xlsx"):
            return {}, [
                ImportIssue(
                    code="INVALID_FILE_TYPE",
                    message="Only .xlsx files are accepted",
                )
            ]
        if len(content) > self.max_size_bytes:
            return {}, [
                ImportIssue(
                    code="FILE_TOO_LARGE",
                    message=(
                        "Workbook exceeds the configured "
                        "size limit"
                    ),
                )
            ]
        try:
            workbook = load_workbook(
                BytesIO(content),
                data_only=False,
                read_only=False,
                keep_vba=False,
            )
        except Exception:
            return {}, [
                ImportIssue(
                    code="INVALID_WORKBOOK",
                    message="Workbook cannot be opened",
                )
            ]

        parsed: dict[str, list[dict[str, Any]]] = {}
        provided_mapping = mapping or {}

        try:
            for sheet_name, columns in SHEET_SPECS.items():
                if sheet_name not in workbook.sheetnames:
                    errors.append(
                        ImportIssue(
                            sheet=sheet_name,
                            code="MISSING_SHEET",
                            message=(
                                "Required worksheet is missing"
                            ),
                        )
                    )
                    continue

                sheet = workbook[sheet_name]
                headers = [
                    cell.value
                    for cell in sheet[1]
                ]
                source_headers = [
                    str(value).strip()
                    for value in headers
                    if value is not None
                    and str(value).strip()
                ]
                allowed_fields = {
                    field
                    for field, _display, _required in columns
                }
                required_fields = {
                    field
                    for field, _display, required in columns
                    if required
                }
                effective_mapping = header_map(sheet_name)
                effective_mapping.update(
                    provided_mapping.get(sheet_name, {})
                )

                invalid_targets = {
                    target
                    for source, target in effective_mapping.items()
                    if source in source_headers
                    and target not in allowed_fields
                }
                if invalid_targets:
                    errors.append(
                        ImportIssue(
                            sheet=sheet_name,
                            code="INVALID_MAPPING",
                            message=(
                                "Mapping contains unsupported "
                                "canonical fields"
                            ),
                        )
                    )
                    parsed[sheet_name] = []
                    continue

                positions = {
                    index: effective_mapping[
                        str(value).strip()
                    ]
                    for index, value in enumerate(headers)
                    if value is not None
                    and str(value).strip()
                    in effective_mapping
                }
                mapped_fields = list(positions.values())
                if len(mapped_fields) != len(
                    set(mapped_fields)
                ):
                    errors.append(
                        ImportIssue(
                            sheet=sheet_name,
                            code="INVALID_MAPPING",
                            message=(
                                "Multiple source headers map "
                                "to the same field"
                            ),
                        )
                    )
                    parsed[sheet_name] = []
                    continue

                missing_fields = (
                    required_fields - set(mapped_fields)
                )
                for field in sorted(missing_fields):
                    errors.append(
                        ImportIssue(
                            sheet=sheet_name,
                            field=field,
                            code="MISSING_HEADER",
                            message=(
                                "Required header is missing"
                            ),
                        )
                    )

                rows: list[dict[str, Any]] = []
                for row_number, row in enumerate(
                    sheet.iter_rows(min_row=2),
                    start=2,
                ):
                    if (
                        row_number - 1
                        > self.max_rows_per_sheet
                    ):
                        errors.append(
                            ImportIssue(
                                sheet=sheet_name,
                                row=row_number,
                                code="ROW_LIMIT_EXCEEDED",
                                message=(
                                    "Worksheet exceeds the "
                                    "configured row limit"
                                ),
                            )
                        )
                        break

                    if any(
                        cell.data_type == "f"
                        for cell in row
                    ):
                        errors.append(
                            ImportIssue(
                                sheet=sheet_name,
                                row=row_number,
                                code="FORMULA_NOT_ALLOWED",
                                message=(
                                    "Formula cells are not "
                                    "allowed"
                                ),
                            )
                        )
                        continue

                    values = {
                        positions[index]: cell.value
                        for index, cell in enumerate(row)
                        if index in positions
                        and cell.value not in (None, "")
                    }
                    if not values:
                        continue

                    operation = normalize_code(
                        values.get("operation")
                        or ImportOperation.UPSERT
                    )
                    if operation not in {
                        item.value
                        for item in ImportOperation
                    }:
                        errors.append(
                            ImportIssue(
                                sheet=sheet_name,
                                row=row_number,
                                field="operation",
                                code="INVALID_OPERATION",
                                message=(
                                    "Operation must be CREATE, "
                                    "UPDATE, or UPSERT"
                                ),
                            )
                        )
                        continue

                    values["operation"] = operation
                    values["_row"] = row_number
                    rows.append(values)

                parsed[sheet_name] = rows
        finally:
            workbook.close()

        return parsed, errors
