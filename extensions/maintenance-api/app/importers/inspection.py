from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import BusinessValidationError
from app.importers.template import SHEET_SPECS, header_map

TEMPLATE_VERSION = "PLAN05-2-TASK09-V1"


def _has_value(row: tuple[Any, ...]) -> bool:
    return any(value not in (None, "") for value in row)


def inspect_workbook(content: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(
            BytesIO(content),
            data_only=False,
            read_only=True,
            keep_vba=False,
        )
    except Exception as exc:
        raise BusinessValidationError(
            "Workbook cannot be opened",
            code="INVALID_WORKBOOK",
        ) from exc

    sheets: list[dict[str, Any]] = []
    sheet_summary: dict[str, int] = {}

    try:
        for sheet_name, columns in SHEET_SPECS.items():
            if sheet_name not in workbook.sheetnames:
                sheets.append(
                    {
                        "name": sheet_name,
                        "source_headers": [],
                        "suggested_mapping": {},
                        "required_fields": [
                            field
                            for field, _display, required in columns
                            if required
                        ],
                        "missing_required_fields": [
                            field
                            for field, _display, required in columns
                            if required
                        ],
                        "row_count": 0,
                    }
                )
                sheet_summary[sheet_name] = 0
                continue

            sheet = workbook[sheet_name]
            header_values = next(
                sheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                ),
                (),
            )
            source_headers = [
                str(value).strip()
                for value in header_values
                if value is not None
                and str(value).strip()
            ]
            suggested_mapping = {
                header: header_map(sheet_name)[header]
                for header in source_headers
                if header in header_map(sheet_name)
            }
            required_fields = [
                field
                for field, _display, required in columns
                if required
            ]
            mapped_fields = set(suggested_mapping.values())
            row_count = sum(
                1
                for row in sheet.iter_rows(
                    min_row=2,
                    values_only=True,
                )
                if _has_value(row)
            )

            sheets.append(
                {
                    "name": sheet_name,
                    "source_headers": source_headers,
                    "suggested_mapping": suggested_mapping,
                    "required_fields": required_fields,
                    "missing_required_fields": [
                        field
                        for field in required_fields
                        if field not in mapped_fields
                    ],
                    "row_count": row_count,
                }
            )
            sheet_summary[sheet_name] = row_count
    finally:
        workbook.close()

    return {
        "template_version": TEMPLATE_VERSION,
        "sheets": sheets,
        "sheet_summary": sheet_summary,
    }
