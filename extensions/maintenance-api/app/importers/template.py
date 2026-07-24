from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_SPECS: dict[str, list[tuple[str, str, bool]]] = {
    "01_装备型号": [
        ("operation", "操作", True),
        ("code", "装备型号编码", True),
        ("name", "装备型号名称", True),
        ("category", "类别", False),
        ("manufacturer", "制造商", False),
        ("model_series", "系列", False),
        ("service_life_years", "设计寿命_年", False),
        ("description", "说明", False),
        ("is_active", "是否启用", False),
    ],
    "02_构型版本": [
        ("operation", "操作", True),
        ("equipment_code", "装备型号编码", True),
        ("version_code", "构型版本编码", True),
        ("version_name", "构型版本名称", True),
        ("effective_date", "生效日期", False),
        ("expiry_date", "失效日期", False),
        ("is_default", "是否默认", False),
        ("source_reference", "来源依据", False),
        ("description", "说明", False),
        ("is_active", "是否启用", False),
    ],
    "03_部件": [
        ("operation", "操作", True),
        ("code", "部件编码", True),
        ("name", "部件名称", True),
        ("part_type", "部件类型", False),
        ("specification", "规格型号", False),
        ("manufacturer", "制造商", False),
        ("unit", "单位", False),
        ("drawing_number", "图号", False),
        ("maintenance_level", "维修等级", False),
        ("description", "说明", False),
        ("is_active", "是否启用", False),
    ],
    "04_维修器材": [
        ("operation", "操作", True),
        ("code", "器材编码", True),
        ("name", "器材名称", True),
        ("specification", "规格型号", False),
        ("category", "类别", False),
        ("unit", "单位", False),
        ("manufacturer", "制造商", False),
        ("material_code", "物料编码", False),
        ("national_standard", "国军标", False),
        ("shelf_life_months", "保质期_月", False),
        ("is_serialized", "是否序列化", False),
        ("is_repairable", "是否可修复", False),
        ("is_critical", "是否关键", False),
        ("default_service_level", "默认服务水平", False),
        ("description", "说明", False),
        ("is_active", "是否启用", False),
    ],
    "05_构型明细": [
        ("operation", "操作", True),
        ("equipment_code", "装备型号编码", True),
        ("version_code", "构型版本编码", True),
        ("item_code", "构型节点编码", True),
        ("parent_item_code", "父节点编码", False),
        ("part_code", "部件编码", True),
        ("spare_part_code", "器材编码", False),
        ("install_quantity", "单机安装数", True),
        ("position_code", "安装位置编码", False),
        ("position_name", "安装位置名称", False),
        ("criticality_level", "关键度", False),
        ("replacement_ratio", "更换比例", False),
        ("maintenance_level", "维修等级", False),
        ("is_mandatory", "是否必装", False),
        ("sort_order", "排序", False),
        ("notes", "备注", False),
    ],
    "06_可靠性参数": [
        ("operation", "操作", True),
        ("profile_code", "参数档案编码", True),
        ("spare_part_code", "器材编码", True),
        ("equipment_code", "装备型号编码", False),
        ("version_code", "构型版本编码", False),
        ("model_type", "模型类型", True),
        ("failure_rate", "失效率", False),
        ("mtbf_hours", "MTBF_小时", False),
        ("weibull_shape", "威布尔形状参数", False),
        ("weibull_scale", "威布尔尺度参数", False),
        ("binomial_trials", "二项试验次数", False),
        ("binomial_probability", "二项概率", False),
        ("negative_binomial_r", "负二项参数r", False),
        ("negative_binomial_p", "负二项概率p", False),
        ("empirical_mean", "经验均值", False),
        ("empirical_variance", "经验方差", False),
        ("extension_parameters_json", "扩展参数JSON", False),
        ("operating_condition_json", "工况JSON", False),
        ("data_source_type", "数据来源类型", True),
        ("data_source_reference", "数据来源说明", False),
        ("sample_size", "样本量", False),
        ("confidence_level", "置信水平", False),
        ("estimated_at", "估计时间", False),
        ("valid_from", "有效起始日期", False),
        ("valid_to", "有效结束日期", False),
        ("notes", "备注", False),
        ("is_active", "是否启用", False),
    ],
    "07_库房": [
        ("operation", "操作", True),
        ("code", "库房编码", True),
        ("name", "库房名称", True),
        ("warehouse_type", "库房类型", False),
        ("location", "位置", False),
        ("organization", "所属单位", False),
        ("responsible_person", "负责人", False),
        ("contact", "联系方式", False),
        ("status", "库房状态", False),
        ("description", "说明", False),
        ("is_active", "是否启用", False),
    ],
    "08_库存": [
        ("operation", "操作", True),
        ("warehouse_code", "库房编码", True),
        ("spare_part_code", "器材编码", True),
        ("on_hand_quantity", "现存数量", True),
        ("reserved_quantity", "预留数量", False),
        ("damaged_quantity", "损坏数量", False),
        ("quarantined_quantity", "隔离数量", False),
        ("in_transit_quantity", "在途数量", False),
        ("safety_stock", "安全库存", False),
        ("reorder_point", "补货点", False),
        ("maximum_stock", "最大库存", False),
        ("last_counted_at", "盘点时间", False),
        ("notes", "备注", False),
    ],
    "09_供应商": [
        ("operation", "操作", True),
        ("code", "供应商编码", True),
        ("name", "供应商名称", True),
        ("supplier_type", "供应商类型", False),
        ("contact_person", "联系人", False),
        ("phone", "电话", False),
        ("email", "邮箱", False),
        ("address", "地址", False),
        ("credit_code", "统一社会信用代码", False),
        ("rating", "评分", False),
        ("qualification_status", "资质状态", False),
        ("description", "说明", False),
        ("is_active", "是否启用", False),
    ],
    "10_供应商报价": [
        ("operation", "操作", True),
        ("offer_code", "报价编码", True),
        ("supplier_code", "供应商编码", True),
        ("spare_part_code", "器材编码", True),
        ("unit_price", "含税或未税单价", True),
        ("currency", "币种", False),
        ("tax_rate", "税率", False),
        ("price_includes_tax", "价格是否含税", False),
        ("lead_time_days", "采购提前期_天", True),
        ("minimum_order_quantity", "最小订购量", False),
        ("order_multiple", "订购批量", False),
        ("maximum_supply_quantity", "最大供应量", False),
        ("warranty_months", "质保期_月", False),
        ("quality_level", "质量等级", False),
        ("is_preferred", "是否首选", False),
        ("valid_from", "报价生效日期", False),
        ("valid_to", "报价失效日期", False),
        ("notes", "备注", False),
        ("is_active", "是否启用", False),
    ],
}


def header_map(sheet_name: str) -> dict[str, str]:
    return {display: field for field, display, _required in SHEET_SPECS[sheet_name]}


def required_headers(sheet_name: str) -> set[str]:
    return {display for _field, display, required in SHEET_SPECS[sheet_name] if required}


def create_template_bytes() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet_name, columns in SHEET_SPECS.items():
        sheet = workbook.create_sheet(sheet_name)
        for index, (_field, display, required) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=index, value=display)
            cell.font = Font(bold=True)
            cell.fill = required_fill if required else header_fill
            sheet.column_dimensions[get_column_letter(index)].width = max(
                14, min(28, len(display) * 2 + 4)
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_template(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(create_template_bytes())
    return path
