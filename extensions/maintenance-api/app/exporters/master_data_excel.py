from __future__ import annotations

from collections.abc import Callable, Mapping
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import or_, select
from sqlalchemy.orm import Session, joinedload

from app.core.exceptions import BusinessValidationError
from app.models.catalog import Part, SparePart
from app.models.equipment import ConfigurationVersion, EquipmentModel
from app.models.inventory import Warehouse
from app.models.reliability import ReliabilityProfile
from app.models.supplier import Supplier, SupplierOffer
from app.security.actor import ActorContext, MaintenanceRole
from app.services.inventory_query_service import inventory_query_service

DEFAULT_MAX_EXPORT_ROWS = 100_000
FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass(frozen=True)
class ExportColumn:
    header: str
    getter: Callable[[Any], object]
    width: int


@dataclass(frozen=True)
class ExportResourceSpec:
    worksheet_title: str
    model: type[Any]
    columns: tuple[ExportColumn, ...]
    keyword_builder: Callable[[str], Any]
    sort_fields: Mapping[str, Any]
    default_sort: str
    filter_builders: Mapping[str, Callable[[object], Any]]
    query_options: tuple[Any, ...] = ()
    related_tenant_builders: tuple[
        Callable[[str], Any],
        ...,
    ] = ()


@dataclass(frozen=True)
class InventoryExportRow:
    warehouse_id: int
    spare_part_id: int
    warehouse_code: str
    spare_part_code: str
    on_hand_quantity: Decimal
    reserved_quantity: Decimal
    damaged_quantity: Decimal
    quarantined_quantity: Decimal
    in_transit_quantity: Decimal
    available_quantity: Decimal
    safety_stock: Decimal
    reorder_point: Decimal
    maximum_stock: Decimal | None
    last_counted_at: datetime | None = None


def _attr(name: str) -> Callable[[Any], object]:
    return lambda row: getattr(row, name)


def _related(
    relationship_name: str,
    attribute_name: str,
) -> Callable[[Any], object]:
    return lambda row: getattr(
        getattr(row, relationship_name),
        attribute_name,
    )


def _column(
    header: str,
    attribute_name: str,
    width: int = 18,
) -> ExportColumn:
    return ExportColumn(
        header=header,
        getter=_attr(attribute_name),
        width=width,
    )


def _related_column(
    header: str,
    relationship_name: str,
    attribute_name: str,
    width: int = 18,
) -> ExportColumn:
    return ExportColumn(
        header=header,
        getter=_related(
            relationship_name,
            attribute_name,
        ),
        width=width,
    )


def _contains(*fields: Any) -> Callable[[str], Any]:
    def build(keyword: str) -> Any:
        pattern = f"%{keyword}%"
        return or_(
            *(
                field.ilike(pattern)
                for field in fields
            )
        )

    return build


def _configuration_keyword(keyword: str) -> Any:
    pattern = f"%{keyword}%"
    return or_(
        ConfigurationVersion.version_code.ilike(
            pattern
        ),
        ConfigurationVersion.version_name.ilike(
            pattern
        ),
        ConfigurationVersion.equipment_model.has(
            or_(
                EquipmentModel.code.ilike(pattern),
                EquipmentModel.name.ilike(pattern),
            )
        ),
    )


def _reliability_keyword(keyword: str) -> Any:
    pattern = f"%{keyword}%"
    return or_(
        ReliabilityProfile.profile_code.ilike(
            pattern
        ),
        ReliabilityProfile.data_source_reference.ilike(
            pattern
        ),
        ReliabilityProfile.spare_part.has(
            or_(
                SparePart.code.ilike(pattern),
                SparePart.name.ilike(pattern),
            )
        ),
    )


def _offer_keyword(keyword: str) -> Any:
    pattern = f"%{keyword}%"
    return or_(
        SupplierOffer.offer_code.ilike(pattern),
        SupplierOffer.supplier.has(
            or_(
                Supplier.code.ilike(pattern),
                Supplier.name.ilike(pattern),
            )
        ),
        SupplierOffer.spare_part.has(
            or_(
                SparePart.code.ilike(pattern),
                SparePart.name.ilike(pattern),
            )
        ),
    )


def _eq(attribute: Any) -> Callable[[object], Any]:
    return lambda value: attribute == value


def _related_tenant(
    relationship: Any,
    related_model: type[Any],
) -> Callable[[str], Any]:
    return lambda tenant_id: relationship.has(
        related_model.tenant_id == tenant_id
    )


RESOURCE_SPECS: Mapping[
    str,
    ExportResourceSpec,
] = {
    "equipment-models": ExportResourceSpec(
        worksheet_title="装备型号",
        model=EquipmentModel,
        columns=(
            _column("编码", "code"),
            _column("名称", "name", 24),
            _column("类别", "category"),
            _column("制造商", "manufacturer", 24),
            _column("系列", "model_series"),
            _column("设计寿命_年", "service_life_years"),
            _column("说明", "description", 32),
            _column("是否启用", "is_active"),
        ),
        keyword_builder=_contains(
            EquipmentModel.code,
            EquipmentModel.name,
            EquipmentModel.category,
            EquipmentModel.manufacturer,
            EquipmentModel.model_series,
        ),
        sort_fields={
            "code": EquipmentModel.code,
            "name": EquipmentModel.name,
            "category": EquipmentModel.category,
            "manufacturer": EquipmentModel.manufacturer,
        },
        default_sort="code",
        filter_builders={
            "is_active": _eq(
                EquipmentModel.is_active
            ),
            "category": _eq(
                EquipmentModel.category
            ),
        },
    ),
    "configuration-versions": ExportResourceSpec(
        worksheet_title="构型版本",
        model=ConfigurationVersion,
        columns=(
            _related_column(
                "装备型号编码",
                "equipment_model",
                "code",
            ),
            _column("构型版本编码", "version_code"),
            _column(
                "构型版本名称",
                "version_name",
                24,
            ),
            _column("状态", "status"),
            _column("生效日期", "effective_date"),
            _column("失效日期", "expiry_date"),
            _column("是否默认", "is_default"),
            _column("说明", "description", 32),
        ),
        keyword_builder=_configuration_keyword,
        sort_fields={
            "version_code": (
                ConfigurationVersion.version_code
            ),
            "version_name": (
                ConfigurationVersion.version_name
            ),
            "status": ConfigurationVersion.status,
            "effective_date": (
                ConfigurationVersion.effective_date
            ),
        },
        default_sort="version_code",
        filter_builders={
            "is_active": _eq(
                ConfigurationVersion.is_active
            ),
            "status": _eq(
                ConfigurationVersion.status
            ),
            "is_default": _eq(
                ConfigurationVersion.is_default
            ),
        },
        query_options=(
            joinedload(
                ConfigurationVersion.equipment_model
            ),
        ),
        related_tenant_builders=(
            _related_tenant(
                ConfigurationVersion.equipment_model,
                EquipmentModel,
            ),
        ),
    ),
    "parts": ExportResourceSpec(
        worksheet_title="部件",
        model=Part,
        columns=(
            _column("部件编码", "code"),
            _column("部件名称", "name", 24),
            _column("部件类型", "part_type"),
            _column("规格型号", "specification", 24),
            _column("制造商", "manufacturer", 24),
            _column("单位", "unit", 12),
            _column("图号", "drawing_number"),
            _column("维修等级", "maintenance_level"),
            _column("说明", "description", 32),
            _column("是否启用", "is_active"),
        ),
        keyword_builder=_contains(
            Part.code,
            Part.name,
            Part.part_type,
            Part.specification,
            Part.manufacturer,
            Part.drawing_number,
        ),
        sort_fields={
            "code": Part.code,
            "name": Part.name,
            "part_type": Part.part_type,
            "manufacturer": Part.manufacturer,
        },
        default_sort="code",
        filter_builders={
            "is_active": _eq(Part.is_active),
            "part_type": _eq(Part.part_type),
        },
    ),
    "spare-parts": ExportResourceSpec(
        worksheet_title="维修器材",
        model=SparePart,
        columns=(
            _column("器材编码", "code"),
            _column("器材名称", "name", 24),
            _column("规格型号", "specification", 24),
            _column("类别", "category"),
            _column("单位", "unit", 12),
            _column("制造商", "manufacturer", 24),
            _column("物料编码", "material_code"),
            _column("国军标", "national_standard"),
            _column("保质期_月", "shelf_life_months"),
            _column("是否序列化", "is_serialized"),
            _column("是否可修复", "is_repairable"),
            _column("是否关键", "is_critical"),
            _column(
                "默认服务水平",
                "default_service_level",
            ),
            _column("说明", "description", 32),
            _column("是否启用", "is_active"),
        ),
        keyword_builder=_contains(
            SparePart.code,
            SparePart.name,
            SparePart.specification,
            SparePart.category,
            SparePart.manufacturer,
            SparePart.material_code,
            SparePart.national_standard,
        ),
        sort_fields={
            "code": SparePart.code,
            "name": SparePart.name,
            "category": SparePart.category,
            "manufacturer": SparePart.manufacturer,
        },
        default_sort="code",
        filter_builders={
            "is_active": _eq(SparePart.is_active),
            "is_critical": _eq(
                SparePart.is_critical
            ),
            "is_serialized": _eq(
                SparePart.is_serialized
            ),
            "is_repairable": _eq(
                SparePart.is_repairable
            ),
            "category": _eq(SparePart.category),
        },
    ),
    "reliability-profiles": ExportResourceSpec(
        worksheet_title="可靠性参数",
        model=ReliabilityProfile,
        columns=(
            _column(
                "参数档案编码",
                "profile_code",
            ),
            _related_column(
                "器材编码",
                "spare_part",
                "code",
            ),
            _column("模型类型", "model_type"),
            _column("失效率", "failure_rate"),
            _column("MTBF_小时", "mtbf_hours"),
            _column("置信水平", "confidence_level"),
            _column(
                "数据来源类型",
                "data_source_type",
            ),
            _column(
                "数据来源说明",
                "data_source_reference",
                28,
            ),
            _column("有效起始日期", "valid_from"),
            _column("有效结束日期", "valid_to"),
            _column("是否启用", "is_active"),
        ),
        keyword_builder=_reliability_keyword,
        sort_fields={
            "profile_code": (
                ReliabilityProfile.profile_code
            ),
            "model_type": (
                ReliabilityProfile.model_type
            ),
            "valid_from": (
                ReliabilityProfile.valid_from
            ),
        },
        default_sort="profile_code",
        filter_builders={
            "is_active": _eq(
                ReliabilityProfile.is_active
            ),
            "model_type": _eq(
                ReliabilityProfile.model_type
            ),
            "data_source_type": _eq(
                ReliabilityProfile.data_source_type
            ),
            "spare_part_id": _eq(
                ReliabilityProfile.spare_part_id
            ),
        },
        query_options=(
            joinedload(
                ReliabilityProfile.spare_part
            ),
        ),
        related_tenant_builders=(
            _related_tenant(
                ReliabilityProfile.spare_part,
                SparePart,
            ),
        ),
    ),
    "warehouses": ExportResourceSpec(
        worksheet_title="库房",
        model=Warehouse,
        columns=(
            _column("库房编码", "code"),
            _column("库房名称", "name", 24),
            _column("库房类型", "warehouse_type"),
            _column("位置", "location", 24),
            _column("所属单位", "organization", 24),
            _column("负责人", "responsible_person"),
            _column("联系方式", "contact"),
            _column("库房状态", "status"),
            _column("是否启用", "is_active"),
        ),
        keyword_builder=_contains(
            Warehouse.code,
            Warehouse.name,
            Warehouse.warehouse_type,
            Warehouse.location,
            Warehouse.organization,
            Warehouse.responsible_person,
        ),
        sort_fields={
            "code": Warehouse.code,
            "name": Warehouse.name,
            "status": Warehouse.status,
        },
        default_sort="code",
        filter_builders={
            "is_active": _eq(Warehouse.is_active),
            "status": _eq(Warehouse.status),
            "warehouse_type": _eq(
                Warehouse.warehouse_type
            ),
        },
    ),
    "suppliers": ExportResourceSpec(
        worksheet_title="供应商",
        model=Supplier,
        columns=(
            _column("供应商编码", "code"),
            _column("供应商名称", "name", 24),
            _column("供应商类型", "supplier_type"),
            _column("联系人", "contact_person"),
            _column("电话", "phone"),
            _column("邮箱", "email", 24),
            _column("地址", "address", 30),
            _column(
                "统一社会信用代码",
                "credit_code",
                22,
            ),
            _column("评分", "rating"),
            _column("资质状态", "qualification_status"),
            _column("是否启用", "is_active"),
        ),
        keyword_builder=_contains(
            Supplier.code,
            Supplier.name,
            Supplier.supplier_type,
            Supplier.contact_person,
            Supplier.phone,
            Supplier.email,
            Supplier.credit_code,
            Supplier.qualification_status,
        ),
        sort_fields={
            "code": Supplier.code,
            "name": Supplier.name,
            "rating": Supplier.rating,
        },
        default_sort="code",
        filter_builders={
            "is_active": _eq(Supplier.is_active),
            "supplier_type": _eq(
                Supplier.supplier_type
            ),
            "qualification_status": _eq(
                Supplier.qualification_status
            ),
        },
    ),
    "supplier-offers": ExportResourceSpec(
        worksheet_title="供应商报价",
        model=SupplierOffer,
        columns=(
            _column("报价编码", "offer_code"),
            _related_column(
                "供应商编码",
                "supplier",
                "code",
            ),
            _related_column(
                "器材编码",
                "spare_part",
                "code",
            ),
            _column("单价", "unit_price"),
            _column("币种", "currency", 10),
            _column("税率", "tax_rate"),
            _column("是否含税", "price_includes_tax"),
            _column("采购提前期_天", "lead_time_days"),
            _column(
                "最小订购量",
                "minimum_order_quantity",
            ),
            _column("订购批量", "order_multiple"),
            _column(
                "最大供应量",
                "maximum_supply_quantity",
            ),
            _column("质保期_月", "warranty_months"),
            _column("质量等级", "quality_level"),
            _column("是否首选", "is_preferred"),
            _column("生效日期", "valid_from"),
            _column("失效日期", "valid_to"),
            _column("是否启用", "is_active"),
        ),
        keyword_builder=_offer_keyword,
        sort_fields={
            "offer_code": SupplierOffer.offer_code,
            "unit_price": SupplierOffer.unit_price,
            "lead_time_days": (
                SupplierOffer.lead_time_days
            ),
            "valid_from": SupplierOffer.valid_from,
        },
        default_sort="offer_code",
        filter_builders={
            "is_active": _eq(
                SupplierOffer.is_active
            ),
            "currency": _eq(
                SupplierOffer.currency
            ),
            "is_preferred": _eq(
                SupplierOffer.is_preferred
            ),
            "spare_part_id": _eq(
                SupplierOffer.spare_part_id
            ),
            "supplier_id": _eq(
                SupplierOffer.supplier_id
            ),
        },
        query_options=(
            joinedload(SupplierOffer.supplier),
            joinedload(SupplierOffer.spare_part),
        ),
        related_tenant_builders=(
            _related_tenant(
                SupplierOffer.supplier,
                Supplier,
            ),
            _related_tenant(
                SupplierOffer.spare_part,
                SparePart,
            ),
        ),
    ),
}


def _safe_excel_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, Enum):
        value = value.value
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.isoformat()
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        stripped = value.lstrip(" \t\r\n")
        if stripped.startswith(FORMULA_PREFIXES):
            return f"'{value}"
    return value


def _inventory_excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return format(value.quantize(Decimal("0.0001")), ".4f")
    return _safe_excel_value(value)


class MasterDataExcelExporter:
    def __init__(
        self,
        *,
        max_rows: int = DEFAULT_MAX_EXPORT_ROWS,
        query_service=inventory_query_service,
    ) -> None:
        self.max_rows = max_rows
        self.query_service = query_service

    def export(
        self,
        session: Session,
        *,
        tenant_id: str,
        resource_key: str,
        filters: Mapping[str, object | None],
    ) -> bytes:
        if resource_key == "inventories":
            return self._export_inventory(
                session,
                tenant_id=tenant_id,
                filters=filters,
            )
        spec = RESOURCE_SPECS.get(resource_key)
        if spec is None:
            raise BusinessValidationError(
                code="EXPORT_RESOURCE_NOT_SUPPORTED",
                message=(
                    "Requested master-data export "
                    "resource is not supported"
                ),
                details={"resource_key": resource_key},
            )

        query = select(spec.model).where(
            spec.model.tenant_id == tenant_id
        )

        if spec.query_options:
            query = query.options(
                *spec.query_options
            )

        for builder in spec.related_tenant_builders:
            query = query.where(
                builder(tenant_id)
            )

        include_inactive = filters.get(
            "include_inactive",
            False,
        )
        active_field = getattr(
            spec.model,
            "is_active",
            None,
        )
        if (
            active_field is not None
            and include_inactive is not True
        ):
            query = query.where(
                active_field.is_(True)
            )

        keyword = filters.get("keyword")
        if keyword is not None:
            normalized = str(keyword).strip()
            if normalized:
                query = query.where(
                    spec.keyword_builder(normalized)
                )

        ignored = {
            "keyword",
            "include_inactive",
            "sort_by",
            "sort_order",
        }
        for name, value in filters.items():
            if name in ignored or value is None:
                continue
            builder = spec.filter_builders.get(name)
            if builder is None:
                raise BusinessValidationError(
                    code="INVALID_EXPORT_FILTER",
                    message=(
                        "Filter is not supported for "
                        "this export resource"
                    ),
                    details={
                        "resource_key": resource_key,
                        "filter": name,
                    },
                )
            query = query.where(builder(value))

        sort_by_value = filters.get(
            "sort_by",
            spec.default_sort,
        )
        sort_by = (
            spec.default_sort
            if sort_by_value is None
            else str(sort_by_value)
        )
        sort_field = spec.sort_fields.get(sort_by)
        if sort_field is None:
            raise BusinessValidationError(
                code="INVALID_EXPORT_SORT_FIELD",
                message=(
                    "Sort field is not supported for "
                    "this export resource"
                ),
                details={
                    "resource_key": resource_key,
                    "sort_by": sort_by,
                },
            )

        sort_order_value = filters.get(
            "sort_order",
            "asc",
        )
        sort_order = (
            "asc"
            if sort_order_value is None
            else str(sort_order_value).lower()
        )
        if sort_order not in {"asc", "desc"}:
            raise BusinessValidationError(
                code="INVALID_EXPORT_SORT_ORDER",
                message=(
                    "Export sort order must be "
                    "'asc' or 'desc'"
                ),
                details={"sort_order": sort_order},
            )

        order_expression = (
            sort_field.asc()
            if sort_order == "asc"
            else sort_field.desc()
        )
        query = query.order_by(
            order_expression,
            spec.model.id.asc(),
        ).limit(self.max_rows + 1)

        rows = list(
            session.execute(query)
            .scalars()
            .unique()
            .all()
        )
        if len(rows) > self.max_rows:
            raise BusinessValidationError(
                code="EXPORT_ROW_LIMIT_EXCEEDED",
                message=(
                    "Export row limit was exceeded"
                ),
                details={
                    "max_rows": self.max_rows,
                    "resource_key": resource_key,
                },
            )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = spec.worksheet_title

        headers = [
            column.header
            for column in spec.columns
        ]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            worksheet.append(
                [
                    _safe_excel_value(
                        column.getter(row)
                    )
                    for column in spec.columns
                ]
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )
        for index, column in enumerate(
            spec.columns,
            start=1,
        ):
            worksheet.column_dimensions[
                worksheet.cell(
                    row=1,
                    column=index,
                ).column_letter
            ].width = column.width

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _export_inventory(
        self,
        session: Session,
        *,
        tenant_id: str,
        filters: Mapping[str, object | None],
    ) -> bytes:
        supported = {
            "keyword",
            "include_inactive",
            "sort_by",
            "sort_order",
            "warehouse_id",
            "spare_part_id",
        }
        unknown = next(
            (
                name
                for name, value in filters.items()
                if name not in supported and value is not None
            ),
            None,
        )
        if unknown is not None:
            raise BusinessValidationError(
                "Filter is not supported for this export resource",
                code="INVALID_EXPORT_FILTER",
                details={"resource_key": "inventories", "filter": unknown},
            )
        actor = ActorContext(
            user_id="master-data-export",
            tenant_id=tenant_id,
            role=MaintenanceRole.VIEWER,
            request_id="master-data-export",
            token_id="master-data-export",
        )
        sort_by = str(filters.get("sort_by") or "last_counted_at")
        if sort_by not in {
            "last_counted_at",
            "on_hand_quantity",
            "available_quantity",
        }:
            raise BusinessValidationError(
                "Sort field is not supported for this export resource",
                code="INVALID_EXPORT_SORT_FIELD",
                details={"resource_key": "inventories", "sort_by": sort_by},
            )
        sort_order = str(filters.get("sort_order") or "asc").lower()
        if sort_order not in {"asc", "desc"}:
            raise BusinessValidationError(
                "Export sort order must be 'asc' or 'desc'",
                code="INVALID_EXPORT_SORT_ORDER",
                details={"sort_order": sort_order},
            )
        rows = [
            InventoryExportRow(**row)
            for row in self.query_service.inventory_export_rows(
                session,
                actor,
                keyword=str(filters.get("keyword") or "").strip(),
                warehouse_id=filters.get("warehouse_id"),
                spare_part_id=filters.get("spare_part_id"),
                sort_by=sort_by,
                sort_order=sort_order,
                limit=self.max_rows + 1,
            )
        ]
        if len(rows) > self.max_rows:
            raise BusinessValidationError(
                "Export row limit was exceeded",
                code="EXPORT_ROW_LIMIT_EXCEEDED",
                details={"max_rows": self.max_rows, "resource_key": "inventories"},
            )
        columns = (
            _column("库房编码", "warehouse_code"),
            _column("器材编码", "spare_part_code"),
            _column("现存数量", "on_hand_quantity"),
            _column("预留数量", "reserved_quantity"),
            _column("损坏数量", "damaged_quantity"),
            _column("隔离数量", "quarantined_quantity"),
            _column("在途数量", "in_transit_quantity"),
            _column("可用数量", "available_quantity"),
            _column("安全库存", "safety_stock"),
            _column("补货点", "reorder_point"),
            _column("最大库存", "maximum_stock"),
            _column("盘点时间", "last_counted_at", 22),
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "库存"
        worksheet.append([column.header for column in columns])
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            worksheet.append(
                [_inventory_excel_value(column.getter(row)) for column in columns]
            )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for index, column in enumerate(columns, start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=index).column_letter
            ].width = column.width
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()
