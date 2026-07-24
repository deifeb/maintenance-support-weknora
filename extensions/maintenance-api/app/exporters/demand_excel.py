from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _safe(value):
    if value is None:
        return ""
    text = str(value.value if hasattr(value, "value") else value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def export_calculation_excel(calculation, runs) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("01_任务摘要")
    summary.append(["字段", "值"])
    for key, value in (
        ("任务编码", calculation.calculation_code),
        ("任务名称", calculation.calculation_name),
        ("状态", calculation.status),
        ("输入快照哈希", calculation.input_snapshot_hash),
    ):
        summary.append([key, _safe(value)])
    result_sheet = wb.create_sheet("02_器材需求结果")
    headers = [
        "运行模式",
        "器材编码",
        "器材名称",
        "期望需求",
        "P50",
        "P80",
        "P90",
        "P95",
        "P99",
        "推荐数量",
        "可用库存",
        "净缺口",
        "风险等级",
    ]
    result_sheet.append(headers)
    for run in runs:
        for item in run.item_results:
            result_sheet.append(
                [
                    run.run_mode.value,
                    _safe(item.spare_part_code_snapshot),
                    _safe(item.spare_part_name_snapshot),
                    float(item.expected_demand),
                    float(item.p50),
                    float(item.p80),
                    float(item.p90),
                    float(item.p95),
                    float(item.p99),
                    float(item.recommended_spare_quantity),
                    float(item.usable_inventory),
                    float(item.net_demand_gap),
                    item.shortage_risk_level.value,
                ]
            )
    for name in (
        "03_阶段汇总",
        "04_需求贡献明细",
        "05_模型与参数快照",
        "06_库存缺口",
        "07_警告与诊断",
    ):
        ws = wb.create_sheet(name)
        ws.append(["说明"])
        ws.append(["数据已包含在任务结果与输入快照中"])
    if len(runs) == 2:
        ws = wb.create_sheet("08_模型对比")
        ws.append(["说明"])
        ws.append(["解析计算与蒙特卡洛结果请通过 comparison API 查看"])
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        ws.freeze_panes = "A2"
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
